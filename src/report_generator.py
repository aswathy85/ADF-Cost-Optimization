"""report_generator.py — Build Excel report workbook and export Power BI CSV.

Contains helper routines for building all report sheets used by the tool.
The main public function is `generate_excel_report(payload, cfg)`.
"""
from __future__ import annotations
import csv, json, logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.config import AppConfig
from src.models import ActivityRecord, FactoryRecord, OptimizationSuggestion, PipelineRecord, ReportPayload, TriggerRecord

logger = logging.getLogger(__name__)

# ── Colour palette ──────────────────────────────────────────────────────────
C = {
    "header_bg": "006E74", "header_fg": "FFFFFF", "subheader": "0097AC",
    "high": "FF6B6B", "medium": "FFD93D", "low": "6BCB77",
    "alt": "EBF3FB", "white": "FFFFFF", "grey": "F2F2F2",
    "border": "B7DDE2", "dark": "231F20",
    "dev": "006E74", "qa": "0097AC", "uat": "6BCB77", "prod": "231F20",
}
PRIORITY_CLR = {"High": "FF6B6B", "Medium": "FFD93D", "Low": "6BCB77"}
TIER_CLR     = {"high": "FF6B6B", "medium": "FFD93D", "low": "6BCB77"}
ENV_CLR      = {"dev": C["dev"], "qa": C["qa"], "uat": C["uat"], "prod": C["prod"]}
ENV_ORDER    = ["dev", "qa", "uat", "prod"]

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, size=10, color="1F1F1F", italic=False): return Font(bold=bold, size=size, color=color, italic=italic)
def _border():
    s = Side(border_style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row, headers, widths=None):
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = _font(bold=True, color=C["header_fg"], size=10)
        cell.fill = _fill(C["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        if widths and ci <= len(widths):
            ws.column_dimensions[get_column_letter(ci)].width = widths[ci-1]

def _data_row(ws, row, values, alt=False):
    bg = C["alt"] if alt else C["white"]
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.fill = _fill(bg)
        cell.border = _border()
        cell.alignment = Alignment(vertical="center")
        if isinstance(v, float):
            cell.number_format = "#,##0.00"

def _add_table(ws, ref, name, style="TableStyleMedium2"):
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if max_row <= min_row:
        return
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(t)

def _has_data_rows(header_row: int, last_row: int) -> bool:
    return last_row > header_row

def _set_auto_filter(ws, ref: str):
    """Set an autofilter only when the range is valid for the current sheet."""
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if max_col >= min_col and max_row >= min_row:
        ws.auto_filter.ref = ref

def _date_label(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return ""

def _add_months(value: datetime, months: int) -> datetime:
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    day = min(value.day, 28)
    return value.replace(year=year, month=month, day=day)

def _selected_env_order(payload) -> List[str]:
    selected = [env for env in getattr(payload, "selected_environments", []) if env]
    if not selected:
        selected = sorted(
            {f.environment for f in payload.factories}
            | {t.environment for t in payload.triggers}
            | set(payload.cost_by_environment.keys()),
            key=lambda env: ENV_ORDER.index(env) if env in ENV_ORDER else 99,
        )
    seen = set()
    return [env for env in selected if not (env in seen or seen.add(env))]

def _title_row(ws, text, cols="A1:L1", size=13):
    ws.merge_cells(cols)
    c = ws["A1"]
    c.value = text
    c.font = _font(bold=True, size=size, color=C["header_fg"])
    c.fill = _fill(C["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False

# ============================================================================
def generate_excel_report(payload: ReportPayload, cfg: AppConfig) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    logger.info("Building Excel report…")
    _sheet_cover(wb, payload)
    _sheet_executive(wb, payload, cfg)
    _sheet_env_compare(wb, payload, cfg)
    _sheet_adf_breakdown(wb, payload, cfg)
    _sheet_pipeline(wb, payload, cfg)
    _sheet_pipeline_runs(wb, payload, cfg)
    _sheet_activity(wb, payload, cfg)
    _sheet_cost_drivers(wb, payload, cfg)
    _sheet_triggers(wb, payload, cfg)
    _sheet_ai_suggestions(wb, payload, cfg)
    _sheet_optimized_code(wb, payload, cfg)
    _sheet_powerbi(wb, payload, cfg)
    out = cfg.output.excel_path
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Saved: %s", out)
    _export_csv(payload, cfg)
    return out

# ── COVER ────────────────────────────────────────────────────────────────────
def _sheet_cover(wb, payload):
    ws = wb.create_sheet("📋 Cover")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 50
    ws.row_dimensions[2].height = 55
    def big(row, col, text, size=22, bold=True, color="1F3864"):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("B2:D2")
    big(2, 2, "ADF Cost Optimization Report", 24)
    ws.cell(row=2, column=2).fill = _fill("EBF3FB")
    big(4, 2, "Azure Data Factory — Multi-Environment Cost Analysis & Gen AI Optimization", 13, False, "2E75B6")
    big(6, 2, f"Generated:       {payload.generated_at.strftime('%Y-%m-%d  %H:%M UTC')}", 11, False, "595959")
    big(7, 2, f"Analysis Period: Last {payload.date_range_days} days", 11, False, "595959")
    big(8, 2, f"Total Estimated Cost: ${payload.total_cost_usd:,.2f} USD", 13, True, "C00000")
    big(10, 2, "Report Contents", 13)
    sheets = [
        ("📊 Executive Summary",       "High-level KPIs, cost totals, top 10 most expensive pipelines"),
        ("🌍 Environment Comparison",   "Side-by-side Dev / QA / UAT / Prod comparison"),
        ("🏭 ADF Factory Breakdown",    "Mid-level: cost per factory across all environments"),
        ("🔗 Pipeline Cost Report",     "Mid-level: cost per pipeline with run metrics"),
        ("⚡ Activity Cost Report",     "Low-level: drilldown to individual activity costs"),
        ("🔍 Cost Drivers Analysis",    "Root-cause analysis: what is driving highest cost"),
        ("🔔 Trigger Status Report",    "Enabled / disabled triggers per environment — unexpected state alerts"),
        ("🤖 AI Optimization",          "Gen AI powered suggestions with estimated savings (USD)"),
        ("💡 Optimized Pipeline Code",  "Ready-to-paste optimized ADF JSON for each pipeline"),
        ("📈 Power BI Data Model",      "Star-schema fact table ready for Power BI import"),
    ]
    for i, (name, desc) in enumerate(sheets):
        ws.cell(row=12+i, column=2, value=name).font = Font(bold=True, size=11, color="1F3864")
        ws.cell(row=12+i, column=3, value=desc).font = Font(size=10, color="595959")

# ── EXECUTIVE SUMMARY ────────────────────────────────────────────────────────
def _sheet_executive(wb, payload, cfg):
    ws = wb.create_sheet("📊 Executive Summary")
    _title_row(ws, "ADF Cost Optimization — Executive Summary")
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = f"Period: Last {payload.date_range_days} days  |  Generated: {payload.generated_at.strftime('%Y-%m-%d %H:%M UTC')}"
    sub.font = _font(italic=True, size=10, color="595959")
    sub.alignment = Alignment(horizontal="center")

    # KPI bar
    ws.row_dimensions[4].height = 22; ws.row_dimensions[5].height = 32
    kpis = [
        ("Total Cost (USD)",   f"${payload.total_cost_usd:,.2f}", "C00000"),
        ("ADF Factories",      str(len(payload.factories)), "2E75B6"),
        ("Total Pipelines",    str(sum(len(f.pipelines) for f in payload.factories)), "ED7D31"),
        ("Enabled Triggers",   str(sum(f.enabled_trigger_count for f in payload.factories)), "A9D18E"),
        ("Environments",       str(len(payload.cost_by_environment)), "1F3864"),
    ]
    for ki, (lbl, val, clr) in enumerate(kpis):
        col = 1 + ki * 2
        lc = ws.cell(row=4, column=col, value=lbl)
        lc.font = _font(bold=True, size=9, color="FFFFFF"); lc.fill = _fill(clr)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=5, column=col, value=val)
        vc.font = _font(bold=True, size=14, color=clr)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions[get_column_letter(col+1)].width = 2

    # Cost by environment table
    ws.cell(row=7, column=1, value="Cost by Environment").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, 8, ["Environment","Cost (USD)","Factories","Pipelines","% of Total"], [16,16,13,13,13])
    env_order = _selected_env_order(payload)
    for i, env in enumerate(env_order):
        cost = payload.cost_by_environment.get(env, 0)
        fcts = [f for f in payload.factories if f.environment == env]
        pls  = sum(len(f.pipelines) for f in fcts)
        pct  = cost / payload.total_cost_usd * 100 if payload.total_cost_usd else 0
        _data_row(ws, 9+i, [env.upper(), round(cost,2), len(fcts), pls, round(pct,1)], alt=i%2==0)
        ec = ws.cell(row=9+i, column=1); ec.fill = _fill(ENV_CLR.get(env,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    _add_table(ws, f"A8:E{8+len(env_order)}", "tbl_exec_env")

    # Bar chart
    if env_order:
        chart = BarChart(); chart.type = "bar"; chart.title = "Cost by Environment"
        chart.x_axis.title = "USD"; chart.y_axis.title = "Environment"; chart.style = 10; chart.width = 16; chart.height = 11
        chart.add_data(Reference(ws, min_col=2, min_row=8, max_row=8+len(env_order)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=9, max_row=8+len(env_order)))
        ws.add_chart(chart, "G7")

    # Top 10 pipelines
    all_pls = sorted([p for f in payload.factories for p in f.pipelines], key=lambda p: p.estimated_cost_usd, reverse=True)[:10]
    r = 15
    ws.cell(row=r, column=1, value="Top 10 Most Expensive Pipelines").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, r+1, ["Pipeline","Factory","Env","Cost (USD)","Runs","Avg Cost/Run","Cost Tier"], [38,32,10,14,10,14,12])
    for i, pl in enumerate(all_pls):
        _data_row(ws, r+2+i, [pl.pipeline_name, pl.factory_name, pl.environment.upper(),
                               round(pl.estimated_cost_usd,2), pl.run_count, round(pl.avg_cost_per_run,4), pl.cost_tier.upper()], alt=i%2==0)
        tc = ws.cell(row=r+2+i, column=7); tc.fill = _fill(TIER_CLR.get(pl.cost_tier, C["white"])); tc.font = _font(bold=True)
        ec = ws.cell(row=r+2+i, column=3); ec.fill = _fill(ENV_CLR.get(pl.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    last = r+1+len(all_pls)
    _add_table(ws, f"A{r+1}:G{last}", "tbl_top10")
    if _has_data_rows(r+1, last):
        ws.conditional_formatting.add(f"D{r+2}:D{last}", DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color="638EC6"))
    _add_scope_and_forecast(ws, payload, start_row=29)

def _add_scope_and_forecast(ws, payload, start_row=29):
    ws.cell(row=start_row, column=1, value="Selected Analysis Scope").font = _font(bold=True, size=12, color=C["header_bg"])
    date_range = f"{_date_label(payload.analysis_start_date)} to {_date_label(payload.analysis_end_date)}"
    if date_range.strip() == "to":
        date_range = f"Last {payload.date_range_days} days"
    envs = _selected_env_order(payload)
    ws.cell(row=start_row + 1, column=1, value="Date Range").font = _font(bold=True)
    ws.cell(row=start_row + 1, column=2, value=date_range)
    ws.cell(row=start_row + 2, column=1, value="Environments").font = _font(bold=True)
    ws.cell(row=start_row + 2, column=2, value=", ".join(e.upper() for e in envs) or "N/A")
    _header_row(ws, start_row + 4, ["Environment", "Subscriptions", "Resource Groups"], [16, 48, 48])
    scope = payload.selected_scope or {}
    if scope:
        for idx, env in enumerate(envs):
            env_scope = scope.get(env, {})
            subs = ", ".join(env_scope.get("subscriptions", [])) or "N/A"
            rgs = ", ".join(env_scope.get("resource_groups", [])) or "N/A"
            _data_row(ws, start_row + 5 + idx, [env.upper(), subs, rgs], alt=idx % 2 == 0)
            ws.cell(row=start_row + 5 + idx, column=1).fill = _fill(ENV_CLR.get(env, "4472C4"))
            ws.cell(row=start_row + 5 + idx, column=1).font = _font(bold=True, color="FFFFFF")

    forecast_col = 7
    ws.cell(row=start_row, column=forecast_col, value="6-Month Estimated Cost Forecast").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, start_row + 1, ["Month", "Actual Cost", "Estimated Cost"], [16, 16, 18])

    forecast_rows = payload.forecast_months
    if not forecast_rows:
        end_dt = payload.analysis_end_date or payload.generated_at
        daily_cost = payload.total_cost_usd / max(payload.date_range_days, 1)
        monthly_estimate = round(daily_cost * 30.4375, 2)
        forecast_rows = [{
            "label": "Selected Range",
            "actual_cost_usd": round(payload.total_cost_usd, 2),
            "estimated_cost_usd": None,
        }] + [
            {"label": _add_months(end_dt, i).strftime("%b %Y"),
             "actual_cost_usd": None,
             "estimated_cost_usd": monthly_estimate}
            for i in range(1, 7)
        ]

    for i, row_data in enumerate(forecast_rows):
        row = start_row + 2 + i
        ws.cell(row=row, column=forecast_col, value=row_data.get("label", ""))
        ws.cell(row=row, column=forecast_col + 1, value=row_data.get("actual_cost_usd"))
        ws.cell(row=row, column=forecast_col + 2, value=row_data.get("estimated_cost_usd"))

    chart = BarChart()
    chart.type = "col"
    chart.title = "Actual vs Estimated Monthly Cost"
    chart.y_axis.title = "USD"
    chart.style = 10
    chart.width = 15
    chart.height = 9
    chart.add_data(Reference(ws, min_col=forecast_col + 1, max_col=forecast_col + 2, min_row=start_row + 1, max_row=start_row + 1 + len(forecast_rows)), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=forecast_col, min_row=start_row + 2, max_row=start_row + 1 + len(forecast_rows)))
    if len(chart.series) >= 2:
        chart.series[0].graphicalProperties.solidFill = "4472C4"
        chart.series[1].graphicalProperties.solidFill = "ED7D31"
    ws.add_chart(chart, f"G{start_row + 10}")

# ── ENVIRONMENT COMPARISON ───────────────────────────────────────────────────
def _sheet_env_compare(wb, payload, cfg):
    ws = wb.create_sheet("🌍 Environment Comparison")
    _title_row(ws, "Environment Comparison — Dev / QA / UAT / Production")
    env_order = _selected_env_order(payload)
    ws["A1"].value = "Environment Comparison - " + (" / ".join(env.upper() for env in env_order) or "Selected Environments")
    if not env_order:
        ws.cell(row=3, column=1, value="No environments were selected for this analysis.")
        return
    row = 3
    for ei, env in enumerate(env_order):
        fcts = [f for f in payload.factories if f.environment == env]
        pls  = [p for f in fcts for p in f.pipelines]
        cost = sum(f.estimated_cost_usd for f in fcts)
        trigs_tot = sum(f.trigger_count for f in fcts)
        trigs_en  = sum(f.enabled_trigger_count for f in fcts)
        color = ENV_CLR.get(env, "4472C4")
        col = 1 + ei * 3
        for co in range(3): ws.column_dimensions[get_column_letter(col+co)].width = 19
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+2)
        hc = ws.cell(row=row, column=col, value=f"  {env.upper()}")
        hc.font = _font(bold=True, size=13, color="FFFFFF"); hc.fill = _fill(color)
        hc.alignment = Alignment(vertical="center"); ws.row_dimensions[row].height = 28
        metrics = [("Total Cost", f"${cost:,.2f}"), ("Factories", str(len(fcts))),
                   ("Pipelines", str(len(pls))), ("Triggers (Total)", str(trigs_tot)),
                   ("Triggers Enabled", str(trigs_en)), ("% Enabled", f"{trigs_en/max(trigs_tot,1)*100:.0f}%")]
        for mi, (lbl, val) in enumerate(metrics):
            lc = ws.cell(row=row+1+mi*2, column=col, value=lbl)
            lc.font = _font(bold=True, size=9, color="595959"); lc.fill = _fill(C["grey"])
            ws.merge_cells(start_row=row+1+mi*2, start_column=col, end_row=row+1+mi*2, end_column=col+2)
            vc = ws.cell(row=row+2+mi*2, column=col, value=val)
            vc.font = _font(bold=True, size=13, color=color)
            ws.merge_cells(start_row=row+2+mi*2, start_column=col, end_row=row+2+mi*2, end_column=col+2)
            ws.row_dimensions[row+2+mi*2].height = 22

    # Comparison table
    rt = row + 17
    ws.cell(row=rt, column=1, value="Detailed Comparison Table").font = _font(bold=True, size=12, color=C["header_bg"])
    rt += 1
    _header_row(ws, rt, ["Metric"]+[e.upper() for e in env_order], [30]+[20]*len(env_order))
    rt += 1
    env_data = {}
    for env in env_order:
        fcts = [f for f in payload.factories if f.environment == env]
        pls  = [p for f in fcts for p in f.pipelines]
        env_data[env] = {
            "Total Cost ($)":        round(sum(f.estimated_cost_usd for f in fcts), 2),
            "Factories":             len(fcts), "Pipelines": len(pls),
            "Pipeline Runs":         sum(p.run_count for p in pls),
            "Triggers (Total)":      sum(f.trigger_count for f in fcts),
            "Triggers (Enabled)":    sum(f.enabled_trigger_count for f in fcts),
            "Avg Cost/Pipeline ($)": round(sum(p.estimated_cost_usd for p in pls)/max(len(pls),1),2),
            "High-Cost Pipelines":   sum(1 for p in pls if p.cost_tier=="high"),
        }
    for mi, metric in enumerate(list(next(iter(env_data.values())).keys())):
        _data_row(ws, rt+mi, [metric]+[env_data[e].get(metric,0) for e in env_order], alt=mi%2==0)
    last = rt + len(env_data[env_order[0]]) - 1
    last_col = get_column_letter(1 + len(env_order))
    _add_table(ws, f"A{rt-1}:{last_col}{last}", "tbl_env_compare")

    # Pie chart
    pie = PieChart(); pie.title = "Cost by Environment"; pie.style = 10; pie.width = 14; pie.height = 10
    pr = last + 3
    ws.cell(row=pr,   column=1, value="Env"); ws.cell(row=pr,   column=2, value="Cost")
    for ei, env in enumerate(env_order):
        ws.cell(row=pr+1+ei, column=1, value=env.upper())
        ws.cell(row=pr+1+ei, column=2, value=env_data[env]["Total Cost ($)"])
    pie.add_data(Reference(ws, min_col=2, min_row=pr, max_row=pr+len(env_order)), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=pr+1, max_row=pr+len(env_order)))
    ws.add_chart(pie, "G" + str(rt-1))

# ── ADF FACTORY BREAKDOWN ────────────────────────────────────────────────────
def _sheet_adf_breakdown(wb, payload, cfg):
    ws = wb.create_sheet("🏭 ADF Factory Breakdown")
    _title_row(ws, "ADF Factory Cost Breakdown — Mid-Level Report")
    headers = ["Factory Name","Environment","Resource Group","Subscription","Location","Est. Cost (USD)","Pipelines","Triggers","Enabled Triggers","% Triggers Active","Cost Tier"]
    widths  = [36,12,28,34,12,16,12,12,16,16,12]
    _header_row(ws, 3, headers, widths)
    fcts = sorted(payload.factories, key=lambda f: f.estimated_cost_usd, reverse=True)
    for i, f in enumerate(fcts):
        pct = f.enabled_trigger_count / max(f.trigger_count,1) * 100
        tier = "HIGH" if f.estimated_cost_usd >= cfg.cost_thresholds.high else "MED" if f.estimated_cost_usd >= cfg.cost_thresholds.medium else "LOW"
        _data_row(ws, 4+i, [f.factory_name, f.environment.upper(), f.resource_group, f.subscription_id, f.location,
                             round(f.estimated_cost_usd,2), f.pipeline_count, f.trigger_count, f.enabled_trigger_count, round(pct,1), tier], alt=i%2==0)
        ec = ws.cell(row=4+i, column=2); ec.fill = _fill(ENV_CLR.get(f.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
        tc = ws.cell(row=4+i, column=11); tc.fill = _fill(TIER_CLR.get(tier.lower(), C["white"])); tc.font = _font(bold=True)
    last = 3 + len(fcts)
    _add_table(ws, f"A3:K{last}", "tbl_adf")
    ws.freeze_panes = "A4"; _set_auto_filter(ws, f"A3:K{last}")
    if _has_data_rows(3, last):
        chart = BarChart(); chart.type = "bar"; chart.title = "Factory Cost Ranking"; chart.style = 10; chart.width = 22; chart.height = 16
        chart.add_data(Reference(ws, min_col=6, min_row=3, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=last))
        ws.add_chart(chart, f"A{last+3}")

# ── PIPELINE COST REPORT ─────────────────────────────────────────────────────
def _sheet_pipeline(wb, payload, cfg):
    ws = wb.create_sheet("🔗 Pipeline Cost Report")
    _title_row(ws, "Pipeline Cost Breakdown — Mid-Level Report")
    headers = ["Pipeline Name","Factory","Environment","Resource Group","Actual Cost (USD)","Est. Cost (USD)","Run Count","Failed Runs","Last Status","Avg Cost/Run (USD)","Avg Duration (min)","Last Run","Cost Tier","Activities"]
    widths  = [40,32,12,28,16,16,12,12,14,18,18,22,12,12]
    _header_row(ws, 3, headers, widths)
    pls = sorted([p for f in payload.factories for p in f.pipelines], key=lambda p: p.actual_cost_usd or p.estimated_cost_usd, reverse=True)
    for i, pl in enumerate(pls):
        lr = pl.last_run_end.strftime("%Y-%m-%d %H:%M") if pl.last_run_end else "N/A"
        _data_row(ws, 4+i, [pl.pipeline_name, pl.factory_name, pl.environment.upper(), pl.resource_group,
                             round(pl.actual_cost_usd,2), round(pl.estimated_cost_usd,2), pl.run_count,
                             pl.failed_run_count, pl.last_run_status or "N/A",
                             round(pl.actual_avg_cost_per_run,4),
                             round(pl.avg_duration_sec/60,1), lr, pl.cost_tier.upper(), len(pl.activities)], alt=i%2==0)
        ec = ws.cell(row=4+i, column=3); ec.fill = _fill(ENV_CLR.get(pl.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
        fc = ws.cell(row=4+i, column=8)
        if pl.failed_run_count:
            fc.fill = _fill("FF6B6B"); fc.font = _font(bold=True)
        sc = ws.cell(row=4+i, column=9)
        if str(pl.last_run_status).lower() == "failed":
            sc.fill = _fill("FF6B6B"); sc.font = _font(bold=True)
        tc = ws.cell(row=4+i, column=13); tc.fill = _fill(TIER_CLR.get(pl.cost_tier, C["white"])); tc.font = _font(bold=True)
        ws.cell(row=4+i, column=5).number_format = "$#,##0.00"
        ws.cell(row=4+i, column=6).number_format = "$#,##0.00"
    last = 3 + len(pls)
    _add_table(ws, f"A3:N{last}", "tbl_pipeline")
    ws.freeze_panes = "A4"; _set_auto_filter(ws, f"A3:N{last}")
    if _has_data_rows(3, last):
        ws.conditional_formatting.add(f"E4:E{last}", DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color="638EC6"))

# ── ACTIVITY COST REPORT ─────────────────────────────────────────────────────
def _sheet_pipeline_runs(wb, payload, cfg):
    ws = wb.create_sheet("Pipeline Run Cost Report")
    _title_row(ws, "Pipeline Run Cost Report - Per Execution")
    headers = ["Run ID","Pipeline","Factory","Environment","Resource Group","Status","Run Start","Run End","Duration (min)","Actual Run Cost (USD)","Estimated Run Cost (USD)"]
    widths = [42,40,32,12,28,14,22,22,16,18,18]
    _header_row(ws, 3, headers, widths)
    rows = []
    for f in payload.factories:
        for p in f.pipelines:
            for run in p.run_costs:
                rows.append((f, p, run))
    rows.sort(key=lambda item: str(item[2].get("run_start") or item[2].get("run_end") or ""), reverse=True)
    for i, (f, p, run) in enumerate(rows):
        start = run.get("run_start")
        end = run.get("run_end")
        start_text = start.strftime("%Y-%m-%d %H:%M") if start else "N/A"
        end_text = end.strftime("%Y-%m-%d %H:%M") if end else "N/A"
        duration_min = float(run.get("duration_ms", 0) or 0) / 60000
        _data_row(ws, 4+i, [
            run.get("run_id", ""), p.pipeline_name, f.factory_name, p.environment.upper(),
            p.resource_group, run.get("status", ""), start_text, end_text,
            round(duration_min, 2), round(float(run.get("actual_cost_usd", 0) or 0), 4),
            round(float(run.get("estimated_cost_usd", 0) or 0), 4)
        ], alt=i%2==0)
        ec = ws.cell(row=4+i, column=4); ec.fill = _fill(ENV_CLR.get(p.environment,"006E74")); ec.font = _font(bold=True, color="FFFFFF")
        status_cell = ws.cell(row=4+i, column=6)
        if str(run.get("status", "")).lower() == "failed":
            status_cell.fill = _fill("FF6B6B"); status_cell.font = _font(bold=True)
    last = 3 + len(rows)
    _add_table(ws, f"A3:K{last}", "tbl_pipeline_runs")
    ws.freeze_panes = "A4"; _set_auto_filter(ws, f"A3:K{last}")
    if _has_data_rows(3, last):
        ws.conditional_formatting.add(f"J4:J{last}", DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color="0097AC"))

def _sheet_activity(wb, payload, cfg):
    ws = wb.create_sheet("⚡ Activity Cost Report")
    _title_row(ws, "Activity Cost Breakdown — Low-Level Drilldown")
    headers = ["Activity Name","Activity Type","Pipeline","Factory","Environment","Run Count","Avg Duration (min)","DIU-Hours","Actual Cost (USD)","Est. Cost (USD)","Cost/Run (USD)","Data Details"]
    widths  = [36,20,32,28,12,12,16,14,16,16,14,28]
    _header_row(ws, 3, headers, widths)
    acts = sorted([a for f in payload.factories for p in f.pipelines for a in p.activities], key=lambda a: a.actual_cost_usd or a.estimated_cost_usd, reverse=True)
    if not acts:
        for f in payload.factories:
            for p in f.pipelines:
                acts.append(ActivityRecord(
                    activity_name=f"Pipeline total: {p.pipeline_name}",
                    activity_type="PipelineAggregate",
                    pipeline_name=p.pipeline_name,
                    factory_name=p.factory_name,
                    environment=p.environment,
                    subscription_id=p.subscription_id,
                    run_count=p.run_count,
                    total_duration_sec=p.total_duration_sec,
                    diu_hours=0.0,
                    estimated_cost_usd=round(p.estimated_cost_usd, 4),
                    actual_cost_usd=round(p.actual_cost_usd or p.estimated_cost_usd, 4),
                    avg_duration_sec=p.avg_duration_sec,
                ))
    for i, a in enumerate(acts):
        cpr = a.actual_cost_usd / max(a.run_count,1)
        details = a.data_details if isinstance(a.data_details, str) else (a.data_details or {})
        if not details and getattr(a, "activity_runs", None):
            status_counts = {}
            for run in a.activity_runs:
                status = run.get("status", "unknown")
                status_counts[status] = status_counts.get(status, 0) + 1
            details = {
                "activity_run_count": len(a.activity_runs),
                "run_status_counts": status_counts,
                "latest_run_start": a.activity_runs[-1].get("run_start"),
                "latest_run_end": a.activity_runs[-1].get("run_end"),
            }
        data_details = details if isinstance(details, str) else json.dumps(details, default=str)
        _data_row(ws, 4+i, [a.activity_name, a.activity_type, a.pipeline_name, a.factory_name,
                             a.environment.upper(), a.run_count, round(a.avg_duration_sec/60,1),
                             round(a.diu_hours,2), round(a.actual_cost_usd,4), round(a.estimated_cost_usd,4),
                             round(cpr,4), data_details], alt=i%2==0)
        ec = ws.cell(row=4+i, column=5); ec.fill = _fill(ENV_CLR.get(a.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    last = 3 + len(acts)
    _add_table(ws, f"A3:L{last}", "tbl_activity")
    ws.freeze_panes = "A4"; _set_auto_filter(ws, f"A3:J{last}")
    if _has_data_rows(3, last):
        ws.conditional_formatting.add(f"I4:I{last}", DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color="FF6B6B"))

# ── COST DRIVERS ─────────────────────────────────────────────────────────────
def _sheet_cost_drivers(wb, payload, cfg):
    ws = wb.create_sheet("🔍 Cost Drivers Analysis")
    _title_row(ws, "Root Cause — Cost Drivers Analysis")
    ws.cell(row=3, column=1, value="Cost by Activity Type").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, 4, ["Activity Type","Total Cost (USD)","Total Runs","Total DIU-Hours","Avg Cost/Run"], [28,18,14,16,16])
    type_costs = {}
    for f in payload.factories:
        for p in f.pipelines:
            for a in p.activities:
                type_costs.setdefault(a.activity_type, {"cost":0,"runs":0,"diu":0})
                type_costs[a.activity_type]["cost"] += a.estimated_cost_usd
                type_costs[a.activity_type]["runs"] += a.run_count
                type_costs[a.activity_type]["diu"]  += a.diu_hours
    sorted_types = sorted(type_costs.items(), key=lambda x: x[1]["cost"], reverse=True)
    for i, (atype, s) in enumerate(sorted_types):
        _data_row(ws, 5+i, [atype, round(s["cost"],2), s["runs"], round(s["diu"],1), round(s["cost"]/max(s["runs"],1),4)], alt=i%2==0)
    last_t = 4 + len(sorted_types)
    _add_table(ws, f"A4:E{last_t}", "tbl_type_cost")

    pie = PieChart(); pie.title = "Cost Share by Activity Type"; pie.style = 10; pie.width = 14; pie.height = 10
    if _has_data_rows(4, last_t):
        pie.add_data(Reference(ws, min_col=2, min_row=4, max_row=last_t), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last_t))
        ws.add_chart(pie, "G3")

    rs = last_t + 3
    ws.cell(row=rs, column=1, value="High-Cost Pipeline Root Cause").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, rs+1, ["Pipeline","Factory","Env","Cost (USD)","Runs","Cost Tier","Primary Cost Driver"], [40,32,10,14,10,12,45])
    all_pls = sorted([p for f in payload.factories for p in f.pipelines], key=lambda p: p.estimated_cost_usd, reverse=True)[:20]
    for i, pl in enumerate(all_pls):
        top_act = max(pl.activities, key=lambda a: a.estimated_cost_usd) if pl.activities else None
        driver  = f"{top_act.activity_type}: {top_act.activity_name} (${top_act.estimated_cost_usd:.2f})" if top_act else "N/A"
        _data_row(ws, rs+2+i, [pl.pipeline_name, pl.factory_name, pl.environment.upper(),
                                round(pl.estimated_cost_usd,2), pl.run_count, pl.cost_tier.upper(), driver], alt=i%2==0)
        tc = ws.cell(row=rs+2+i, column=6); tc.fill = _fill(TIER_CLR.get(pl.cost_tier, C["white"])); tc.font = _font(bold=True)
        ec = ws.cell(row=rs+2+i, column=3); ec.fill = _fill(ENV_CLR.get(pl.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    last2 = rs + 1 + len(all_pls)
    _add_table(ws, f"A{rs+1}:G{last2}", "tbl_drivers")
    ws.freeze_panes = f"A{rs+2}"; _set_auto_filter(ws, f"A{rs+1}:G{last2}")

# ── TRIGGER STATUS ───────────────────────────────────────────────────────────
def _sheet_triggers(wb, payload, cfg):
    ws = wb.create_sheet("🔔 Trigger Status Report")
    _title_row(ws, "ADF Trigger Status — All Environments")
    ws.cell(row=3, column=1, value="Summary by Environment").font = _font(bold=True, size=11, color=C["header_bg"])
    _header_row(ws, 4, ["Environment","Total Triggers","Enabled","Disabled","% Enabled","⚠️ Alert"], [16,16,12,12,12,35])
    env_order = _selected_env_order(payload)
    for ei, env in enumerate(env_order):
        ets     = [t for t in payload.triggers if t.environment == env]
        enabled = sum(1 for t in ets if t.is_enabled)
        pct     = enabled / max(len(ets),1) * 100
        flag    = ""
        if env in ("dev","qa") and pct > 75: flag = "⚠️ High trigger rate in non-prod!"
        elif env == "prod" and pct < 50:     flag = "⚠️ Low trigger activity in prod!"
        _data_row(ws, 5+ei, [env.upper(), len(ets), enabled, len(ets)-enabled, round(pct,1), flag], alt=ei%2==0)
        ec = ws.cell(row=5+ei, column=1); ec.fill = _fill(ENV_CLR.get(env,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    summary_last = 4 + len(env_order)
    _add_table(ws, f"A4:F{summary_last}", "tbl_trig_summary")

    rt = 11
    ws.cell(row=rt, column=1, value="All Triggers — Detailed").font = _font(bold=True, size=11, color=C["header_bg"])
    headers = ["Trigger Name","Factory","Environment","Resource Group","Type","Status","Enabled?","Schedule","Pipelines Triggered","Last Triggered"]
    widths  = [40,32,12,28,28,12,10,28,45,22]
    _header_row(ws, rt+1, headers, widths); rt += 2
    trigs = sorted(payload.triggers, key=lambda t: (["prod","uat","qa","dev"].index(t.environment) if t.environment in ["prod","uat","qa","dev"] else 99, t.factory_name))
    for i, tr in enumerate(trigs):
        lt  = tr.last_triggered.strftime("%Y-%m-%d %H:%M") if tr.last_triggered else "Never"
        pls = ", ".join(tr.pipelines_triggered) if tr.pipelines_triggered else "—"
        _data_row(ws, rt+i, [tr.trigger_name, tr.factory_name, tr.environment.upper(), tr.resource_group,
                              tr.trigger_type, tr.status, "✅ Yes" if tr.is_enabled else "❌ No",
                              tr.schedule_expression, pls, lt], alt=i%2==0)
        ec = ws.cell(row=rt+i, column=3); ec.fill = _fill(ENV_CLR.get(tr.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
        sc = ws.cell(row=rt+i, column=7); sc.fill = _fill("6BCB77" if tr.is_enabled else "FF6B6B"); sc.font = _font(bold=True)
    last = rt + len(trigs) - 1
    _add_table(ws, f"A{rt-1}:J{last}", "tbl_triggers")
    ws.freeze_panes = f"A{rt}"; _set_auto_filter(ws, f"A{rt-1}:J{last}")

# ── AI OPTIMIZATION ──────────────────────────────────────────────────────────
def _sheet_ai_suggestions(wb, payload, cfg):
    ws = wb.create_sheet("🤖 AI Optimization")
    _title_row(ws, "🤖 Gen AI Powered — Pipeline Cost Optimization Suggestions")
    sugs = payload.optimization_suggestions
    if not sugs:
        ws.cell(row=3, column=1, value="No AI suggestions — set ANTHROPIC_API_KEY in .env and re-run.")
        return
    total_save  = sum(s.estimated_saving_usd for s in sugs)
    high_pri    = sum(1 for s in sugs if s.priority == "High")
    unique_pls  = len(set(s.pipeline_name for s in sugs))
    kpis = [("Total Potential Savings", f"${total_save:,.2f}", "C00000"),
            ("Pipelines Analysed", str(unique_pls), "2E75B6"),
            ("High Priority Issues", str(high_pri), "ED7D31"),
            ("Total Suggestions", str(len(sugs)), "A9D18E")]
    ws.row_dimensions[3].height = 22; ws.row_dimensions[4].height = 32
    for ki, (lbl, val, clr) in enumerate(kpis):
        col = 1 + ki * 3
        lc = ws.cell(row=3, column=col, value=lbl); lc.font = _font(bold=True, size=9, color="FFFFFF"); lc.fill = _fill(clr); lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row=4, column=col, value=val); vc.font = _font(bold=True, size=13, color=clr); vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 24
        ws.column_dimensions[get_column_letter(col+1)].width = 24
        ws.column_dimensions[get_column_letter(col+2)].width = 3

    rs = 6
    ws.cell(row=rs, column=1, value="Optimization Recommendations").font = _font(bold=True, size=12, color=C["header_bg"])
    _header_row(ws, rs+1, ["Pipeline","Factory","Environment","Current Cost ($)","Issue Category","Issue Description","Recommendation","Saving (%)","Saving (USD)","Priority","Effort"],
                [38,30,12,14,28,55,65,12,14,12,12])
    sugs_s = sorted(sugs, key=lambda s: s.estimated_saving_usd, reverse=True)
    for i, s in enumerate(sugs_s):
        _data_row(ws, rs+2+i, [s.pipeline_name, s.factory_name, s.environment.upper(),
                                round(s.current_cost_usd,2), s.issue_category, s.issue_description,
                                s.suggestion, round(s.estimated_saving_pct,1), round(s.estimated_saving_usd,2),
                                s.priority, s.effort], alt=i%2==0)
        ws.row_dimensions[rs+2+i].height = 55
        for col_w in [6,7]:
            ws.cell(row=rs+2+i, column=col_w).alignment = Alignment(wrap_text=True, vertical="top")
        pc = ws.cell(row=rs+2+i, column=10); pc.fill = _fill(PRIORITY_CLR.get(s.priority, C["white"])); pc.font = _font(bold=True)
        ec = ws.cell(row=rs+2+i, column=3); ec.fill = _fill(ENV_CLR.get(s.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
    last = rs+1+len(sugs_s)
    _add_table(ws, f"A{rs+1}:K{last}", "tbl_ai_sugs")
    ws.freeze_panes = f"A{rs+2}"; _set_auto_filter(ws, f"A{rs+1}:K{last}")

    # Savings bar chart
    ps: Dict[str, float] = {}
    for s in sugs: ps[f"{s.pipeline_name}({s.environment})"] = ps.get(f"{s.pipeline_name}({s.environment})",0) + s.estimated_saving_usd
    cr = last + 3
    ws.cell(row=cr, column=1, value="Pipeline"); ws.cell(row=cr, column=2, value="Savings ($)")
    for ci, (pl, sv) in enumerate(sorted(ps.items(), key=lambda x: -x[1])[:15]):
        ws.cell(row=cr+1+ci, column=1, value=pl); ws.cell(row=cr+1+ci, column=2, value=round(sv,2))
    chart = BarChart(); chart.type = "bar"; chart.title = "Potential Savings by Pipeline (USD)"; chart.style = 10; chart.width = 22; chart.height = 14
    if ps:
        chart.add_data(Reference(ws, min_col=2, min_row=cr, max_row=cr+min(len(ps),15)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=cr+1, max_row=cr+min(len(ps),15)))
        ws.add_chart(chart, f"L{rs+1}")

# ── OPTIMIZED CODE ───────────────────────────────────────────────────────────
def _sheet_optimized_code(wb, payload, cfg):
    ws = wb.create_sheet("💡 Optimized Pipeline Code")
    _title_row(ws, "Optimized ADF Pipeline Code — Ready to Paste")
    ws.merge_cells("A2:F2")
    ws["A2"].value = ("HOW TO USE: Copy the 'Optimized Code Snippet' column value. "
                      "In ADF Portal → open the Pipeline → click the Activity → switch to 'Code View' → paste the updated properties.")
    ws["A2"].font = _font(italic=True, size=9, color="595959"); ws["A2"].alignment = Alignment(wrap_text=True); ws.row_dimensions[2].height = 28
    sugs = payload.optimization_suggestions
    if not sugs:
        ws.cell(row=4, column=1, value="No suggestions — run with ANTHROPIC_API_KEY set."); return
    widths = [38,30,12,28,55,60]
    _header_row(ws, 4, ["Pipeline","Factory","Env","Issue Category","Optimized Code Snippet","What Changed / Why"], widths)
    seen = set()
    cr = 5
    for s in sorted(sugs, key=lambda x: x.estimated_saving_usd, reverse=True):
        key = (s.pipeline_name, s.factory_name, s.issue_category)
        if key in seen: continue
        seen.add(key)
        snippet = s.optimized_code_snippet or "# See AI Optimization sheet for details"
        _data_row(ws, cr, [s.pipeline_name, s.factory_name, s.environment.upper(), s.issue_category, snippet, s.suggestion[:220]])
        ws.row_dimensions[cr].height = max(60, len(snippet)//2)
        code_c = ws.cell(row=cr, column=5)
        code_c.alignment = Alignment(wrap_text=True, vertical="top"); code_c.font = Font(name="Courier New", size=8); code_c.fill = _fill("F2F2F2")
        desc_c = ws.cell(row=cr, column=6); desc_c.alignment = Alignment(wrap_text=True, vertical="top")
        ec = ws.cell(row=cr, column=3); ec.fill = _fill(ENV_CLR.get(s.environment,"4472C4")); ec.font = _font(bold=True, color="FFFFFF")
        cr += 1

# ── POWER BI DATA MODEL ──────────────────────────────────────────────────────
def _sheet_powerbi(wb, payload, cfg):
    ws = wb.create_sheet("📈 Power BI Data Model")
    _title_row(ws, "Power BI Data Model — Fact Table (Import into Power BI Desktop)")
    ws.merge_cells("A2:N2")
    ws["A2"].value = "Power BI Desktop: Get Data → Excel Workbook → select this file → load sheet '📈 Power BI Data Model'. Build hierarchy: Environment → Factory → Pipeline → Activity"
    ws["A2"].font = _font(italic=True, size=9, color="595959"); ws["A2"].alignment = Alignment(horizontal="left")
    headers = ["FactKey","Date","SubscriptionId","Environment","FactoryName","ResourceGroup","PipelineName","ActivityName","ActivityType","RunCount","AvgDurationMin","DIUHours","CostUSD","CostTier"]
    widths  = [40,14,34,12,34,26,40,38,25,12,16,14,14,12]
    _header_row(ws, 4, headers, widths)
    row = 5
    date_str = payload.generated_at.strftime("%Y-%m-%d")
    for f in payload.factories:
        for p in f.pipelines:
            for a in p.activities:
                _data_row(ws, row, [f"{f.factory_name}|{p.pipeline_name}|{a.activity_name}", date_str,
                                    f.subscription_id, f.environment, f.factory_name, f.resource_group,
                                    p.pipeline_name, a.activity_name, a.activity_type, a.run_count,
                                    round(a.avg_duration_sec/60,2), round(a.diu_hours,4), round(a.estimated_cost_usd,4), p.cost_tier], alt=row%2==0)
                row += 1
    last = row - 1
    if last >= 5: _add_table(ws, f"A4:N{last}", "tbl_fact"); ws.freeze_panes = "A5"; _set_auto_filter(ws, f"A4:N{last}")

    # AI suggestions for Power BI
    sr = last + 4
    ws.cell(row=sr, column=1, value="AI Suggestions Dimension (for Power BI slicers)").font = _font(bold=True, size=11, color=C["header_bg"])
    _header_row(ws, sr+1, ["Pipeline","Factory","Environment","CurrentCostUSD","IssueCategory","SavingPct","SavingUSD","Priority","Effort"],
                [38,30,12,16,28,14,14,12,12])
    for i, s in enumerate(payload.optimization_suggestions):
        _data_row(ws, sr+2+i, [s.pipeline_name, s.factory_name, s.environment, round(s.current_cost_usd,2),
                                s.issue_category, round(s.estimated_saving_pct,1), round(s.estimated_saving_usd,2), s.priority, s.effort])

def _export_csv(payload, cfg):
    path = cfg.output.powerbi_csv_path
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    date_str = payload.generated_at.strftime("%Y-%m-%d")
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["FactKey","Date","SubscriptionId","Environment","FactoryName","ResourceGroup","PipelineName","ActivityName","ActivityType","RunCount","AvgDurationMin","DIUHours","CostUSD","CostTier"])
        for f in payload.factories:
            for p in f.pipelines:
                for a in p.activities:
                    w.writerow([f"{f.factory_name}|{p.pipeline_name}|{a.activity_name}", date_str,
                                f.subscription_id, f.environment, f.factory_name, f.resource_group,
                                p.pipeline_name, a.activity_name, a.activity_type, a.run_count,
                                round(a.avg_duration_sec/60,2), round(a.diu_hours,4), round(a.estimated_cost_usd,4), p.cost_tier])
    logger.info("Power BI CSV: %s", path)
